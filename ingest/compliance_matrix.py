"""Build a bid-desk compliance workbook from match results (Phase 11 Lite).

One harvested requirement → one Compliance row. Suggests C/A from MSI
support_level; never auto-writes N. Stakeholder columns left blank for humans.
"""
from __future__ import annotations

import hashlib
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

SUPPORT_RANK = {"native": 3, "option": 2, "partner": 1}
CONFIDENCE_FLOOR = 0.75

COMPLIANCE_FIELDS = [
    "req_id",
    "source_file",
    "page",
    "requirement",
    "compliance_code",
    "response_notes",
    "exception_or_alternate",
    "mapped",
    "capability_id",
    "capability_alias",
    "capability_name",
    "confidence",
    "match_method",
    "msi_products",
    "support_level",
    "owner",
    "due_date",
    "status",
    "risk",
    "comments",
]


def _req_id(requirement: str, page: Any, index: int) -> str:
    digest = hashlib.sha1(
        f"{page or ''}|{requirement}".encode("utf-8", errors="ignore")
    ).hexdigest()[:10]
    page_part = f"P{page}" if page is not None else "P0"
    return f"{page_part}-{index:04d}-{digest}"


def _best_support(msi: list[dict]) -> str:
    best = ""
    best_rank = -1
    for row in msi:
        level = (row.get("support_level") or "").strip().lower()
        rank = SUPPORT_RANK.get(level, 0)
        if rank > best_rank:
            best_rank = rank
            best = level
    return best


def _msi_products_str(msi: list[dict]) -> str:
    parts = []
    for x in msi:
        name = x.get("product_name") or x.get("family") or x.get("product_id") or ""
        level = x.get("support_level") or ""
        if name or level:
            parts.append(f"{name}:{level}".strip(":"))
    return "; ".join(parts)


def suggest_compliance_code(
    *,
    mapped: bool,
    confidence: float | None,
    support_level: str,
) -> str:
    """Suggest C/A only when MSI evidence is strong; never suggest N."""
    if not mapped:
        return ""
    if confidence is None or float(confidence) < CONFIDENCE_FLOOR:
        return ""
    level = (support_level or "").lower()
    if level == "native":
        return "C"
    if level in ("option", "partner"):
        return "A"
    return ""


def build_compliance_rows(
    results: list[dict],
    *,
    source_file: str = "",
) -> list[dict]:
    rows: list[dict] = []
    for i, r in enumerate(results, start=1):
        matches = r.get("matches") or []
        m = matches[0] if matches else {}
        msi = r.get("msi_coverage") or []
        if not isinstance(msi, list):
            msi = []
        mapped = not r.get("unmapped")
        confidence = m.get("confidence")
        support = _best_support(msi)
        code = suggest_compliance_code(
            mapped=mapped,
            confidence=confidence if confidence is not None else None,
            support_level=support,
        )
        risk = ""
        if not mapped:
            risk = "High"
        elif confidence is not None and float(confidence) < CONFIDENCE_FLOOR:
            risk = "Medium"
        rows.append(
            {
                "req_id": _req_id(r.get("requirement") or "", r.get("page"), i),
                "source_file": source_file,
                "page": r.get("page") if r.get("page") is not None else "",
                "requirement": r.get("requirement") or "",
                "compliance_code": code,
                "response_notes": "",
                "exception_or_alternate": "",
                "mapped": mapped,
                "capability_id": m.get("capability_id") or "",
                "capability_alias": m.get("capability_alias") or "",
                "capability_name": m.get("capability_name") or "",
                "confidence": confidence if confidence is not None else "",
                "match_method": m.get("method") or "",
                "msi_products": _msi_products_str(msi),
                "support_level": support,
                "owner": "",
                "due_date": "",
                "status": "Draft",
                "risk": risk,
                "comments": "",
            }
        )
    return rows


def summarize_rows(rows: list[dict], *, source_file: str = "") -> dict:
    n = len(rows)
    mapped = sum(1 for r in rows if r.get("mapped"))
    suggested_c = sum(1 for r in rows if r.get("compliance_code") == "C")
    suggested_a = sum(1 for r in rows if r.get("compliance_code") == "A")
    blank_code = sum(1 for r in rows if not r.get("compliance_code"))
    return {
        "source_file": source_file,
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "requirements": n,
        "mapped": mapped,
        "unmapped": n - mapped,
        "map_rate": round(mapped / n, 3) if n else 0.0,
        "suggested_c": suggested_c,
        "suggested_a": suggested_a,
        "blank_for_review": blank_code,
    }


def write_compliance_xlsx(
    path: Path,
    rows: list[dict],
    *,
    source_file: str = "",
) -> dict:
    """Write Compliance / Summary / Gaps sheets. Returns summary dict."""
    from openpyxl import Workbook

    summary = summarize_rows(rows, source_file=source_file)
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance"
    ws.append(COMPLIANCE_FIELDS)
    for row in rows:
        ws.append([row.get(f, "") for f in COMPLIANCE_FIELDS])

    ws_sum = wb.create_sheet("Summary")
    for key, val in summary.items():
        ws_sum.append([key, val])

    gaps = [
        r
        for r in rows
        if (not r.get("mapped"))
        or (
            r.get("confidence") != ""
            and r.get("confidence") is not None
            and float(r["confidence"]) < CONFIDENCE_FLOOR
        )
    ]
    ws_gaps = wb.create_sheet("Gaps")
    ws_gaps.append(COMPLIANCE_FIELDS)
    for row in gaps:
        ws_gaps.append([row.get(f, "") for f in COMPLIANCE_FIELDS])

    wb.save(path)
    return summary


def build_workbook_bytes(
    results: list[dict],
    *,
    source_file: str = "",
) -> tuple[bytes, dict]:
    """In-memory xlsx for API download."""
    import io

    from openpyxl import Workbook

    rows = build_compliance_rows(results, source_file=source_file)
    summary = summarize_rows(rows, source_file=source_file)

    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance"
    ws.append(COMPLIANCE_FIELDS)
    for row in rows:
        ws.append([row.get(f, "") for f in COMPLIANCE_FIELDS])

    ws_sum = wb.create_sheet("Summary")
    for key, val in summary.items():
        ws_sum.append([key, val])

    gaps = [
        r
        for r in rows
        if (not r.get("mapped"))
        or (
            r.get("confidence") != ""
            and r.get("confidence") is not None
            and float(r["confidence"]) < CONFIDENCE_FLOOR
        )
    ]
    ws_gaps = wb.create_sheet("Gaps")
    ws_gaps.append(COMPLIANCE_FIELDS)
    for row in gaps:
        ws_gaps.append([row.get(f, "") for f in COMPLIANCE_FIELDS])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), summary
